from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("sample1.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

ws.column_dimensions["A"].width = 5

ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True) # strike : 취소선
c1.font = Font(color="0000FF", size=20, underline="single")

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점이상 효과
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center") # 중앙정렬

        if cell.column ==1:
            continue
        
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            cell.font = Font(color="FF0000")

# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정

wb.save("sample_style.xlsx")