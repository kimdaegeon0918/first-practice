from openpyxl import Workbook
wb = Workbook()
ws = wb.active

title = ["학번","출석","퀴즈1","퀴즈2","중간고사","기말고사","프로젝트","총점","성적"]

# title 입력
for x in range(1,10):
    ws.cell(row=1, column=x, value=title[x-1])

data = [1,10,8,5,14,26,12,
2,7,3,7,15,24,18,
3,9,5,8,8,12,4,
4,7,8,7,17,21,18,
5,7,8,7,16,25,15,
6,3,5,8,8,17,0,
7,4,9,10,16,27,18,
8,6,6,6,15,19,17,
9,10,10,9,19,30,19,
10,9,8,8,20,25,20]

# data, 총점 계산 입력
index = 0
for row in range(2,12):
    sum = 0
    for col in range(1,8):
        if col == 4:
            ws.cell(row=row, column=col, value=10)
            sum += 10
        else:
            ws.cell(row=row, column=col, value=data[index])
            if col != 1 :
                sum += data[index]
        index += 1
    ws.cell(row=row, column=8, value=sum)   

# 성적 입력
for row in range(2,12):
    if int(ws["H{}".format(row)].value) >= 90:
        ws["I{}".format(row)] = "A"
    elif int(ws["H{}".format(row)].value) >= 80:
        ws["I{}".format(row)] = "B"
    elif int(ws["H{}".format(row)].value) >= 70:
        ws["I{}".format(row)] = "C"
    else:
        ws["I{}".format(row)] = "D"

    if int(ws["B{}".format(row)].value) < 5:
        ws["I{}".format(row)] = "F"

wb.save("quiz.xlsx")