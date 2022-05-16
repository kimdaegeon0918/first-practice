from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"].value)  # 1
print(ws["A10"].value) # None

print(ws.cell(row=1, column=1).value) # ws["A1"].value

c = ws.cell(column=3, row=1, value=10) # ws["C1"].value = 10
print(c.value)  # 10

from random import *

for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=x, column=y, value=randint(0,100))

wb.save("sample.xlsx")