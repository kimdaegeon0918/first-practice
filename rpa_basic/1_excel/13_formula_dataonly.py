from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져오기
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 실제 데이터를 가지고옴
# evaluate 되지 않은 상태의 데이터는 None라고 표기됨 - 파일열고 저장후 실행
for row in ws.values:
     for cell in row:
         print(cell)